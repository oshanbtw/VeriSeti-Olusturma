import requests
from bs4 import BeautifulSoup
from openpyxl import *

#Siteye bot olmadığımızı gösteren adres diyelim.
header_param = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.66 Safari/537.36 Edg/87.0.664.41"}
#Veri çekeceğimiz siteler.
url = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50"
url2 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=2"
url3 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=3"
url4 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=4"
url5 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=5"
url6 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=6"
url7 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=7"
url8 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=8"
url9 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=9"
url10 = "https://www.arabam.com/ikinci-el/otomobil/fiat-linea-1-3-multijet-active-plus?take=50&page=10"

#Excel sayfamızı oluşturduk, başlıkları yerleştirdik ve "veriseti" adında py dosyasının olduğu yere kaydettik.
kitap = Workbook()
sheet = kitap.active
sheet.append(("Araç Rengi","Araç Yılı","Araç KM","Araç Fiyatı"))



#Veri çekme işlemi
dikey = 2
sayac = 0
while (sayac < 10):
    #Siteyi alıyoruz.
    if sayac == 0:
        r = requests.get(url, headers = header_param)
    elif sayac == 1:
        r = requests.get(url2, headers = header_param)
    elif sayac == 2:
        r = requests.get(url3, headers = header_param)
    elif sayac == 3:
        r = requests.get(url4, headers = header_param)
    elif sayac == 4:
        r = requests.get(url5, headers = header_param)
    elif sayac == 5:
        r = requests.get(url6, headers = header_param)
    elif sayac == 6:
        r = requests.get(url7, headers = header_param)
    elif sayac == 7:
        r = requests.get(url8, headers = header_param)
    elif sayac == 8:
        r = requests.get(url9, headers = header_param)
    elif sayac == 9:
        r = requests.get(url10, headers = header_param)
    #print(r.status_code) eğer 200 yazdırırsa başarılı
    sayac += 1
    soup = BeautifulSoup(r.content, "html.parser") #sayfanın kaynak kodunu "soup" değişkenine aktarıyoruz.
    
    gelenVeri = soup.find_all("table",{"class":"table listing-table w100 border-grey2"}) #İlk parçalama işlemimiz.
    
    #print(gelenVeri[0].contents)
    #print(len(gelenVeri[0].contents))
    
    aracTablosu = (gelenVeri[0].contents) [len(gelenVeri[0].contents)-1] #İkinci parçalama işlemimiz.
    
    #print(aracTablosu)
    
    aracTablosu = aracTablosu.find_all("tr",{"class":"listing-list-item pr should-hover bg-white"}) #Üçüncü parçalama işlemimiz.
    

    yatay = 1
    for arac in aracTablosu:
        aracBilgileri = arac.find_all("td", {"class":"listing-text pl8 pr8 tac pr"}) #Araç yılı, km ve rengi için parçalama.
        aracFiyatlari = arac.find_all("td", {"class":"pl8 pr8 tac pr"}) # Araç fiyatı için parçalama.
        #Parçalanmış verileri çekiyoruz.
        aracYili = aracBilgileri[0].text
        aracKM = aracBilgileri[1].text
        aracKM = aracKM.replace(".","")
        aracRenk = aracBilgileri[2].text
        aracRenk = aracRenk.replace("-","")
        aracFiyat = aracFiyatlari[0].text
        aracFiyat = aracFiyat.replace("TL","")
        aracFiyat = aracFiyat.replace(" ","")
        aracFiyat = aracFiyat.replace(".","")
        if len(aracFiyat) > 7:
            aracFiyat = aracFiyat[5:10]
        
        #Excel tablomuza gönderiyoruz.
        sheet.cell(row = dikey, column = yatay, value = aracRenk)
        yatay += 1
        sheet.cell(row = dikey, column = yatay, value = aracYili)
        yatay += 1
        sheet.cell(row = dikey, column = yatay, value = aracKM)
        yatay += 1
        sheet.cell(row = dikey, column = yatay, value = aracFiyat)
        dikey += 1
        yatay = 1
    
    
kitap.save("veriseti.xlsx")
kitap.close() #Üstteki satırda kaydediyoruz ve bu satırda kapatıyoruz.

    

